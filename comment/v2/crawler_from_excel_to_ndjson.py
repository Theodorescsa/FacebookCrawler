# -*- coding: utf-8 -*-
import os
import sys
import time
import json
import hashlib
import zipfile
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

# ===== project sys.path =====
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# ===== project imports =====
from get_comment_fb_utils import open_reel_comments_if_present
from get_comment_fb_automation import (
    install_early_hook,
    hook_graphql,
)
import get_comment_fb_automation as _core

from util.startdriverproxy import bootstrap_auth, start_driver_with_proxy
from logs.loging_config import logger

# ---------------------------
# Helpers: ENV + Args
# ---------------------------
def _env(key: str, default: Optional[str] = None, cast=str):
    v = os.environ.get(key, default)
    if v is None:
        return None
    if cast is bool:
        return str(v).lower() in ("1", "true", "yes", "y", "on")
    try:
        return cast(v)
    except Exception:
        return default

def _bool_env(key: str, default: bool = True) -> bool:
    return _env(key, str(default), cast=bool)

# ---------------------------
# Paths compute (comments – per sheet)
# ---------------------------
def compute_paths(
    data_root: Path,
    page_name: str,
    sheet_tag: str,
):
    """
    Trả về các đường dẫn chuẩn:
    - INPUT_EXCEL (mặc định: database/post/page/<page>/thoibao-de-last-split-<sheet_tag>.xlsx)
    - OUTPUT_NDJSON_DIR: database/comment/page/<page>/<sheet_tag>/ndjson_per_post
    - ERROR_EXCEL:       database/comment/page/<page>/<sheet_tag>/crawl_errors-<sheet_tag>.xlsx
    - STATUS_STORE_PATH: database/comment/page/<page>/<sheet_tag>/status_store_<sheet_tag>.json
    - TMP_DIR:           database/comment/page/<page>/<sheet_tag>/tmp_comments_<sheet_tag>
    - DEDUP_CACHE_PATH:  database/comment/page/<page>/<sheet_tag>/reply_dedup_cache_<sheet_tag>.json
    - RAW_DUMPS_DIR:     database/comment/page/<page>/<sheet_tag>/raw_dump_comments
    """
    page_dir = data_root / "comment" / "page" / page_name / sheet_tag
    raw_dumps = page_dir / "raw_dump_comments"
    out_ndjson = page_dir / "ndjson_per_post"
    error_xlsx = page_dir / f"crawl_errors-{sheet_tag}.xlsx"
    status_json = page_dir / f"status_store_{sheet_tag}.json"
    tmp_dir = page_dir / f"tmp_comments_{sheet_tag}"
    dedup_cache = page_dir / f"reply_dedup_cache_{sheet_tag}.json"

    # input excel mặc định (có thể override bằng ENV INPUT_EXCEL)
    post_dir = data_root / "post" / "page" / page_name
    default_in = post_dir / f"thoibao-de-last-split-{sheet_tag}.xlsx"

    # ensure dirs
    for p in [page_dir, raw_dumps, out_ndjson, tmp_dir]:
        p.mkdir(parents=True, exist_ok=True)

    return {
        "INPUT_EXCEL_DEFAULT": default_in,
        "OUTPUT_NDJSON_DIR": out_ndjson,
        "ERROR_EXCEL": error_xlsx,
        "STATUS_STORE_PATH": status_json,
        "TMP_DIR": tmp_dir,
        "DEDUP_CACHE_PATH": dedup_cache,
        "RAW_DUMPS_DIR": raw_dumps,
    }

# =========================
# NDJSON Helpers (per-post)
# =========================
def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def _hash_postlink(postlink: str) -> str:
    return hashlib.md5(postlink.encode("utf-8")).hexdigest()[:16]

def ndjson_path_for_post(output_ndjson_dir: Path, postlink: str) -> str:
    _ensure_dir(str(output_ndjson_dir))
    return str(output_ndjson_dir / f"comments_{_hash_postlink(postlink)}.ndjson")

def append_ndjson_lines_atomic(path: str, items: list[dict]):
    """
    Ghi NDJSON theo lô, atomic-ish:
    - Ghi vào .part trước, sau đó append vào file chính (hạn chế half-line nếu crash).
    """
    _ensure_dir(str(Path(path).parent))
    tmp = f"{path}.part"
    with open(tmp, "w", encoding="utf-8") as f:
        for it in items:
            f.write(json.dumps(it, ensure_ascii=False))
            f.write("\n")
    with open(path, "a", encoding="utf-8") as out, open(tmp, "r", encoding="utf-8") as part:
        for line in part:
            out.write(line)
    try:
        os.remove(tmp)
    except Exception as e:
        logger.error(f"Failed to remove {tmp}: {e}")

def read_existing_pairs_in_file(path: str) -> set[tuple[str, str]]:
    """
    Dedupe theo từng file post:
    - Trả về set((postlink, id_like)) đã tồn tại trong file NDJSON của post đó.
    """
    pairs = set()
    if not os.path.exists(path):
        return pairs
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            try:
                obj = json.loads(s)
            except Exception:
                continue
            pl = str(obj.get("postlink") or "")
            cid = str(
                obj.get("id")
                or obj.get("raw_comment_id")
                or obj.get("reply_id")
                or obj.get("legacy_fbid")
                or ""
            )
            if pl and cid:
                pairs.add((pl, cid))
    return pairs

# =========================
# Excel Error Log
# =========================
def _is_valid_xlsx(path: str) -> bool:
    if not os.path.exists(path) or os.path.getsize(path) < 100:
        return False
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            return '[Content_Types].xml' in zf.namelist()
    except Exception:
        return False

def _atomic_save_wb(wb, path: str):
    tmp = f"{path}.tmp"
    wb.save(tmp)
    os.replace(tmp, path)

def ensure_error_excel(path: str):
    if not _is_valid_xlsx(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["link", "error"])
        _atomic_save_wb(wb, path)

def append_error(path: str, link: str, error: str):
    ensure_error_excel(path)
    try:
        wb = load_workbook(path)
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(["link", "error"])
        _atomic_save_wb(wb, path)
        wb = load_workbook(path)
    ws = wb.active
    ws.append([link, error])
    _atomic_save_wb(wb, path)

# =========================
# Status JSON helpers
# =========================
def _atomic_write_json(path: str, data: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def load_status_store(path: str) -> dict[str, str]:
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f) or {}
                return {str(k): str(v) for k, v in d.items()}
        except Exception:
            logger.error(f"Function: load_status_store - ❌ Error loading status: {path}")
    return {}

def save_status_store(path: str, store: dict[str, str]):
    try:
        _atomic_write_json(path, store)
        logger.info(f"Function: save_status_store - ✅ Saved status: {len(store)} items → {path}")
    except Exception as e:
        logger.error(f"Function: save_status_store - ❌ Error saving status: {e}")
        raise

def get_status(store: dict[str, str], postlink: str) -> str:
    return (store.get(postlink) or "").strip().lower()

def set_status(store: dict[str, str], postlink: str, status: str):
    store[postlink] = status

# =========================
# Patch chặn replies lặp (KHÔNG sửa core)
# =========================
def _load_cache(path: Path):
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f) or {}
        except Exception as e:
            logger.error(f"Function: _load_cache - Failed to load dedup cache: {e}")
    return {}

def _save_cache(path: Path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# will set later after paths resolved
_reply_cache = None
__orig_crawl_replies = _core.crawl_replies_for_parent_expansion

def _patched_crawl_replies_for_parent_expansion(
    driver,
    url,
    form,
    base_reply_vars,
    parent_id,
    parent_token,
    out_json,
    extract_fn,
    clean_fn,
    max_reply_pages=None
):
    global _reply_cache
    post_key = str(getattr(driver, "current_url", "") or url)
    seen_for_post = _reply_cache.setdefault(post_key, {})

    if parent_id in seen_for_post:
        logger.info(f"Function: _patched_crawl_replies_for_parent_expansion - [PATCH] Skip replies for parent={str(parent_id)[:12]}… (already processed for this post)")
        return

    __orig_crawl_replies(
        driver,
        url,
        form,
        base_reply_vars,
        parent_id,
        parent_token,
        out_json,
        extract_fn,
        clean_fn,
        max_reply_pages=max_reply_pages
    )

    seen_for_post[parent_id] = int(time.time())
    # path for cache will be injected from main after env resolved
    # (_dedup_cache_path) is available via closure or passed; we’ll save in main flow
    # we just leave in-memory update here.

# =========================
# Helpers: NDJSON per-post temp + loader
# =========================
def build_post_temp_paths(tmp_dir: Path, postlink: str) -> tuple[str, str]:
    """Sinh đường dẫn out_json & checkpoint riêng theo postlink (hash)"""
    h = hashlib.md5(postlink.encode("utf-8")).hexdigest()[:16]
    out_json = str(tmp_dir / f"comments_{h}.ndjson")
    ckpt    = str(tmp_dir / f"checkpoint_{h}.json")
    return out_json, ckpt

def load_ndjson(path: str) -> List[Dict[str, Any]]:
    out = []
    if not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            try:
                out.append(json.loads(s))
            except Exception:
                pass
    return out

# =========================
# Blocked-in-country precheck
# =========================
class BlockedInCountryError(Exception):
    pass

def _is_blocked_in_country(driver, timeout_sec: float = 2.5) -> str | None:
    import time as _t
    needles = [
        "Bài viết này không hiển thị tại Việt Nam",
        "Do chúng tôi đáp ứng yêu cầu từ Vietnam Ministry of Culture, Sports and Tourism",
        "This content isn't available in your country",
        "This post isn't available in your country",
        "not available in your country",
    ]
    t0 = _t.time()
    while _t.time() - t0 < timeout_sec:
        src = (driver.page_source or "").lower()
        for n in needles:
            if n.lower() in src:
                return n
        _t.sleep(0.2)
    return None

# =========================
# Crawl 1 post
# =========================
def crawl_one_post(driver, postlink: str,raw_dump_path, tmp_dir: Path, max_pages=None) -> List[Dict[str, Any]]:
    out_json, ckpt = build_post_temp_paths(tmp_dir, postlink)
    # clear tạm (nếu muốn resume thì comment 2 dòng dưới)
    if os.path.exists(out_json):
        os.remove(out_json)
    if os.path.exists(ckpt):
        os.remove(ckpt)

    # mở link & precheck
    driver.get(postlink)

    time.sleep(0.8)
    reason = _is_blocked_in_country(driver, timeout_sec=2.5)
    if reason:
        logger.error(f"Function: crawl_one_post - blocked_in_vietnam: {reason}")
        raise BlockedInCountryError(f"blocked_in_vietnam: {reason}")

    time.sleep(0.4)
    hook_graphql(driver)
    time.sleep(0.4)

    if "reel" in postlink:
        try:
            open_reel_comments_if_present(driver)
        except Exception:
            pass

    # gọi core (append vào ndjson tạm)
    _ = _core.crawl_comments(
        driver,
        raw_dump_path,
        out_json=out_json,
        checkpoint_path=ckpt,
        max_pages=max_pages
    )
    # đọc lại ndjson đã append
    return load_ndjson(out_json)

# =========================
# Main crawl loop (Excel sheet) — NDJSON per-post
# =========================
def crawl_from_excel_stream(
    input_path: str,
    sheet_name: str,
    raw_dump_path: str,
    output_ndjson_dir: Path,
    status_store_path: Path,
    error_path: Path,
    tmp_dir: Path,
    dedup_cache_path: Path,
    driver,
    max_retries: int = 0,
):
    # Load status store JSON (link -> status)
    status_store = load_status_store(str(status_store_path))

    # set up reply cache (global) + patch core
    global _reply_cache
    _reply_cache = _load_cache(dedup_cache_path)
    _core.crawl_replies_for_parent_expansion = _patched_crawl_replies_for_parent_expansion

    # đọc sheet (pandas)
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    total = len(df)

    logger.info(f"▶️ Bắt đầu crawl (sheet={sheet_name}), tổng {total} dòng")
    for i in range(total):
        postlink = str(df.iloc[i].get("link") or "").strip()
        if not postlink:
            continue

        current_status = get_status(status_store, postlink)
        if current_status == "done":
            logger.info(f"⏩ [{i+1}/{total}] SKIP (done): {postlink}")
            continue

        logger.info(f"=== [{i+1}/{total}] Crawl: {postlink}")
        success = False
        last_error = None

        for attempt in range(max_retries + 1):
            try:
                records = crawl_one_post(driver, postlink,raw_dump_path, tmp_dir=tmp_dir, max_pages=None)

                # print vài reply mẫu để verify
                replies = [r for r in records if r.get("is_reply")]
                logger.info(f"[VERIFY] replies fetched: {len(replies)}")

                # file đích cho post này
                post_file = ndjson_path_for_post(output_ndjson_dir, postlink)

                # dedupe theo post-file
                existing_pairs = read_existing_pairs_in_file(post_file)

                # === GHI THEO LÔ → NDJSON/POST ===
                batch_items = []
                batch_keys = []

                for c in records:
                    # 1) id dùng để dedupe
                    cid = (
                        c.get("id")
                        or c.get("raw_comment_id")
                        or c.get("reply_id")
                        or c.get("legacy_fbid")
                        or None
                    )
                    if not cid:
                        base = "|".join([
                            postlink or "",
                            str(c.get("parent_id") or ""),
                            str(c.get("created_time") or ""),
                            (str(c.get("content") or c.get("text") or c.get("message") or ""))[:80],
                        ])
                        cid = hashlib.md5(base.encode("utf-8")).hexdigest()

                    key = (postlink, cid)
                    if key in existing_pairs:
                        continue

                    # 2) giữ nguyên payload, thêm 'postlink' nếu chưa có
                    item = dict(c)
                    if "postlink" not in item:
                        item["postlink"] = postlink

                    batch_items.append(item)
                    batch_keys.append(key)

                new_cnt = 0
                if batch_items:
                    append_ndjson_lines_atomic(post_file, batch_items)
                    for k in batch_keys:
                        existing_pairs.add(k)
                    new_cnt = len(batch_items)

                replies_cnt = sum(1 for r in batch_items if r.get("is_reply"))
                logger.info(f"[WRITE:NDJSON] +{new_cnt} rows (replies={replies_cnt}) → {post_file}")

                # save dedup cache after each success
                _save_cache(dedup_cache_path, _reply_cache)

                success = True
                break

            except BlockedInCountryError as e:
                last_error = str(e)
                logger.error(f"[BLOCKED] {postlink} → {last_error}")
                break

            except Exception as e:
                last_error = str(e)
                logger.error(f"[WARN] crawl fail {postlink} (attempt {attempt+1}/{max_retries+1}): {e}")
                time.sleep(1)

        # cập nhật trạng thái
        if success:
            set_status(status_store, postlink, "done")
        else:
            append_error(str(error_path), postlink, last_error or "unknown error")
            set_status(status_store, postlink, "fail")
            logger.info(f"[SKIP] bỏ qua bài: {postlink}")

        # save dần
        save_status_store(str(status_store_path), status_store)

    logger.info(f"✅ DONE sheet {sheet_name} — NDJSON folder: {output_ndjson_dir} — errors: {error_path}")
    save_status_store(str(status_store_path), status_store)

# =========================
# RUN
# =========================
if __name__ == "__main__":
    import argparse
    def env_default(key, default=None, cast=str):
        v = os.environ.get(key)
        if v is None:
            return default
        if cast is bool:
            return str(v).lower() in ("1", "true", "yes", "y", "on")
        try:
            return cast(v)
        except Exception:
            return default

    parser = argparse.ArgumentParser("comment-crawler (trimmed)")
    # bắt buộc / quan trọng
    parser.add_argument("--page-name",     default=env_default("PAGE_NAME", "thoibaode"))
    parser.add_argument("--sheet-name",    default=env_default("SHEET_NAME", "Sheet_3"))

    # đường dẫn dữ liệu cơ bản (nếu không truyền sẽ dùng biến cứng ở đầu file)
    parser.add_argument("--cookies-path",  default=env_default("COOKIE_PATH", None))

    # mitm & proxy (proxy optional)
    parser.add_argument("--mitm-port",     type=int, default=env_default("MITM_PORT", 8899, int))
    parser.add_argument("--proxy-host",    default=env_default("PROXY_HOST", None))
    parser.add_argument("--proxy-port",    type=int, default=env_default("PROXY_PORT", None, int))
    parser.add_argument("--proxy-user",    default=env_default("PROXY_USER", None))
    parser.add_argument("--proxy-pass",    default=env_default("PROXY_PASS", None))

    # headless + retries
    headless_default = env_default("HEADLESS", True, bool)
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--headless", dest="headless", action="store_true", default=headless_default)
    group.add_argument("--no-headless", dest="headless", action="store_false")

    parser.add_argument("--max-retries", type=int, default=env_default("MAX_RETRIES", 0, int))

    # optional override của các file chính (nếu muốn ghi đè)
    parser.add_argument("--input-excel",  default=env_default("INPUT_EXCEL", None))

    args = parser.parse_args()

    # ---- Resolve cơ bản (fallback về các biến đã khai báo ở đầu file) ----
    PAGE_NAME   = args.page_name
    SHEET_NAME  = args.sheet_name

    # base DB path (same as trước)
    DATABASE_PATH = Path(__file__).resolve().parent.parent.parent / "database"

    # cookie default (cậu có thể override bằng --cookies-path)
    DEFAULT_COOKIE_PATH = Path(r"E:\NCS\fb-selenium\database\facebookaccount\authen_tranhoangdinhnam\cookies.json")

    # build paths (only use args.input_excel when provided)
    if args.input_excel:
        INPUT_EXCEL = Path(args.input_excel)
    else:
        # tên file giống cấu trúc cậu dùng: thoibao-de-last-split-Sheet_3.xlsx
        INPUT_EXCEL = DATABASE_PATH / "post" / "page" / PAGE_NAME / f"thoibao-de-last-split-{SHEET_NAME}.xlsx"

    # NOTE: use f-strings so SHEET_NAME được chèn vào path
    OUTPUT_NDJSON_DIR = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / "ndjson_per_post")
    ERROR_EXCEL       = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / f"crawl_errors-{SHEET_NAME}.xlsx")
    STATUS_STORE_PATH = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / f"status_store_{SHEET_NAME}.json")
    TMP_DIR           = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / f"tmp_comments_{SHEET_NAME}")
    DEDUP_CACHE_PATH  = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / f"reply_dedup_cache_{SHEET_NAME}.json")
    RAW_DUMPS         = Path(DATABASE_PATH / "comment" / "page" / PAGE_NAME / f"{SHEET_NAME}" / f"raw_dump_comments_{SHEET_NAME}")

    COOKIE_PATH = Path(args.cookies_path) if args.cookies_path else DEFAULT_COOKIE_PATH
    MITM_PORT   = int(args.mitm_port)
    HEADLESS    = bool(args.headless)

    PROXY_HOST  = args.proxy_host
    PROXY_PORT  = args.proxy_port
    PROXY_USER  = args.proxy_user
    PROXY_PASS  = args.proxy_pass

    # đảm bảo folder tồn tại
    for d in [OUTPUT_NDJSON_DIR, TMP_DIR, RAW_DUMPS]:
        try:
            d.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

    logger.info(f"[BOOT] PAGE={PAGE_NAME} | SHEET={SHEET_NAME} | MITM={MITM_PORT} | DATABASE={DATABASE_PATH}")
    logger.info(f"[PATH] INPUT={INPUT_EXCEL} | OUT_DIR={OUTPUT_NDJSON_DIR} | STATUS={STATUS_STORE_PATH} | ERR={ERROR_EXCEL}")

    # ---- start mitm + chrome ----
    d = start_driver_with_proxy(
        proxy_host=PROXY_HOST,
        proxy_port=PROXY_PORT,
        proxy_user=PROXY_USER,
        proxy_pass=PROXY_PASS,
        mitm_port=MITM_PORT,
        headless=HEADLESS,
    )
    d.set_script_timeout(40)
    try:
        d.execute_cdp_cmd("Network.enable", {})
        d.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled": True})
    except Exception:
        pass

    # auth bằng cookie nếu có
    if COOKIE_PATH and COOKIE_PATH.exists():
        bootstrap_auth(d, str(COOKIE_PATH))
    else:
        logger.warning(f"[AUTH] Cookie path not found or not provided: {COOKIE_PATH}")

    # cài hook sớm
    install_early_hook(d)

    # chạy chính
    crawl_from_excel_stream(
        INPUT_EXCEL,
        SHEET_NAME,
        RAW_DUMPS,
        OUTPUT_NDJSON_DIR,
        STATUS_STORE_PATH,
        ERROR_EXCEL,
        TMP_DIR,
        DEDUP_CACHE_PATH,
        driver=d,
        max_retries=args.max_retries,
    )
    # d.quit()
