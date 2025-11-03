# rebuild_comments_excel_fast.py
import os, re, json, hashlib, shutil, tempfile
from pathlib import Path
from typing import Any, Dict, Iterable, List
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

from openpyxl import load_workbook, Workbook
from openpyxl.workbook import Workbook as WB
from openpyxl.writer.excel import save_workbook
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl import Workbook

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl import Workbook

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook as _Workbook
from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook as _WB
from openpyxl.workbook import Workbook

from openpyxl import Workbook as _W
from openpyxl import load_workbook

# ========= CONFIG =========
INPUT_DIR      = r"E:\NCS\fb-selenium\database\comment\page\thoibaode\sheet1\tmp_comments_sheet1"
OUTPUT_EXCEL   = r"E:\NCS\fb-selenium\database\comment\page\thoibaode\sheet1\thoibaode-comments-sheet1-newv2.xlsx"
REBUILD_EXCEL  = True   
ALLOWED_EXT    = {".json", ".ndjson"}
BATCH_SIZE     = 5000   

COLUMNS = [
    "id","type","postlink","commentlink","author_id","author","author_link","avatar",
    "created_time","content","image_url","like","comment","haha","wow","sad","love",
    "angry","care","video","source_id","is_share","link_share","type_share"
]

# ========= UTILS =========
_SPLIT_JSON_RE = re.compile(r'(?<=\})\s*(?=\{)')

def _safe_json_loads(s: str):
    s = s.strip()
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, list) else [obj]
    except Exception:
        parts = _SPLIT_JSON_RE.split(s)
        out = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            try:
                out.append(json.loads(p))
            except Exception:
                pass
        return out

def _iter_records_from_file(path: Path):
    try:
        if path.suffix.lower() == ".ndjson":
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        yield json.loads(line)
                    except Exception:
                        continue
        else:
            with open(path, "r", encoding="utf-8") as f:
                raw = f.read()
            for obj in _safe_json_loads(raw):
                if isinstance(obj, dict) and ("id" in obj or "raw_comment_id" in obj or "reply_id" in obj):
                    yield obj
                elif isinstance(obj, list):
                    for it in obj:
                        if isinstance(it, dict):
                            yield it
    except Exception:
        return

def _strip_comment_id(qs_url: str) -> str:
    try:
        u = urlsplit(qs_url)
        qs = dict(parse_qsl(u.query, keep_blank_values=True))
        for k in ("comment_id", "reply_comment_id", "m_entstream_source"):
            qs.pop(k, None)
        new_q = urlencode(qs, doseq=True)
        return urlunsplit((u.scheme, u.netloc, u.path, new_q, u.fragment)).rstrip("?")
    except Exception:
        return qs_url

def _ensure_postlink(rec: Dict) -> str:
    postlink = (rec.get("postlink") or "").strip()
    if postlink:
        return postlink
    commentlink = (rec.get("commentlink") or rec.get("link") or "").strip()
    return _strip_comment_id(commentlink) if commentlink else ""

def _jsonify_if_needed(val):
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return val

def _normalize(rec: Dict) -> Dict:
    out = {
        "id": rec.get("id") or rec.get("raw_comment_id") or rec.get("reply_id"),
        "type": rec.get("type") or ("Reply" if rec.get("is_reply") else "Comment"),
        "postlink": _ensure_postlink(rec),
        "commentlink": rec.get("commentlink") or rec.get("link"),
        "author_id": rec.get("author_id"),
        "author": rec.get("author"),
        "author_link": rec.get("author_link"),
        "avatar": _jsonify_if_needed(rec.get("avatar")),
        "created_time": rec.get("created_time"),
        "content": rec.get("content") or rec.get("text") or rec.get("message")
                   or (rec.get("body") if isinstance(rec.get("body"), str) else None),
        "image_url": _jsonify_if_needed(rec.get("image_url")),
        "like": rec.get("like", 0),
        "comment": rec.get("comment", 0),
        "haha": rec.get("haha", 0),
        "wow": rec.get("wow", 0),
        "sad": rec.get("sad", 0),
        "love": rec.get("love", 0),
        "angry": rec.get("angry", 0),
        "care": rec.get("care", 0),
        "video": _jsonify_if_needed(rec.get("video")),
        "source_id": rec.get("source_id"),
        "is_share": rec.get("is_share", False),
        "link_share": rec.get("link_share"),
        "type_share": rec.get("type_share") or "shared_none",
    }
    out["id"] = str(out["id"]).strip() if out["id"] else ""
    out["postlink"] = str(out["postlink"]).strip() if out["postlink"] else ""
    return out

def _read_existing_pairs(path: str) -> set[tuple[str, str]]:
    pairs = set()
    if not os.path.exists(path):
        return pairs
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_id = header.index("id")
        idx_pl = header.index("postlink")
        for row in ws.iter_rows(min_row=2, values_only=True):
            pl = str(row[idx_pl] or "").strip()
            cid = str(row[idx_id] or "").strip()
            if pl and cid:
                pairs.add((pl, cid))
        wb.close()
    except Exception:
        pass
    return pairs

def _iter_existing_rows(path: str):
    """Stream toàn bộ file cũ (kể cả header) → yield list cell values.
       Dùng cho mode append (nhưng thực chất là rebuild streaming)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)
    wb.close()

def _write_streaming_xlsx(output_path: str, rows_iterable: Iterable[List[Any]]):
    """Ghi XLSX ở chế độ streaming (WriteOnly)."""
    from openpyxl import Workbook as WriteOnlyWorkbook
    wb = WriteOnlyWorkbook(write_only=True)
    ws = wb.create_sheet()
    # write
    cnt = 0
    buffer = []
    for row in rows_iterable:
        buffer.append(row)
        if len(buffer) >= BATCH_SIZE:
            for r in buffer:
                ws.append(r)
            buffer.clear()
        cnt += 1
    if buffer:
        for r in buffer:
            ws.append(r)
    # save once
    wb.save(output_path)

def main():
    in_dir = Path(INPUT_DIR)
    files = [p for p in in_dir.glob("**/*") if p.suffix.lower() in ALLOWED_EXT]
    # Nếu không cần thứ tự, bỏ sort() để tiết kiệm
    # files.sort()

    print(f"[SCAN] Found {len(files)} files")

    # 1) dedup set từ file cũ (nếu tồn tại)
    exist_pairs = _read_existing_pairs(OUTPUT_EXCEL)
    print(f"[INIT] Existing pairs: {len(exist_pairs)}")

    # 2) gom tất cả rows (cả cũ lẫn mới) theo dạng streaming:
    rows_to_write = []

    def _yield_all_rows():
        # header
        yield COLUMNS

        # nếu file cũ tồn tại & REBUILD_EXCEL=False: stream copy lại toàn bộ
        # (Thực tế: luôn rebuild file mới -> nhanh & an toàn)
        if os.path.exists(OUTPUT_EXCEL) and not REBUILD_EXCEL:
            for r in _iter_existing_rows(OUTPUT_EXCEL):
                # đảm bảo header đúng format (nếu tên cột khác nhau)
                if r and r[0] == "id":
                    continue  # bỏ header cũ, đã ghi header mới ở trên
                yield r

        # 3) duyệt nguồn dữ liệu, chuẩn hoá & chống trùng, yield row mới
        total_seen, count_new = 0, 0
        for fp in files:
            for rec in _iter_records_from_file(fp):
                total_seen += 1
                norm = _normalize(rec)
                if not norm["id"] or not norm["postlink"]:
                    continue
                key = (norm["postlink"], norm["id"])
                if key in exist_pairs:
                    continue
                exist_pairs.add(key)
                count_new += 1
                yield [norm.get(c) for c in COLUMNS]
        print(f"[DONE] New rows: {count_new}. Total scanned: {total_seen}")

    # 4) ghi theo streaming vào file tạm, rồi replace
    with tempfile.TemporaryDirectory() as td:
        tmp_out = os.path.join(td, "tmp.xlsx")
        _write_streaming_xlsx(tmp_out, _yield_all_rows())
        os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
        shutil.move(tmp_out, OUTPUT_EXCEL)

    print(f"[OK] Output → {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()
